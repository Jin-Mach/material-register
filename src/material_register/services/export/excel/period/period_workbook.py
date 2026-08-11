from pathlib import Path

from openpyxl import Workbook

from material_register.domain.export_dataclass import ExportItemIn, ExportItemOut
from material_register.services.export.excel.period.period_sheet import PeriodSheet


class PeriodWorkbook:
    ERROR_TEXT = "[N/A]"

    @staticmethod
    def create_workbook(
        export_settings: dict[str, Path | str | float | bool],
        export_texts: dict[str, dict[str, str]],
        in_data: list[ExportItemIn],
        out_data: list[ExportItemOut],
    ) -> tuple[Workbook, float]:
        workbook = Workbook()
        workbook.remove(workbook.active)
        period_texts = export_texts.get("PeriodSheet", {})
        sheet = workbook.create_sheet(
            period_texts.get("sheetName", PeriodWorkbook.ERROR_TEXT)
        )
        _, last_balance = PeriodSheet.create_sheet(
            sheet, export_settings, period_texts, in_data, out_data
        )
        return workbook, last_balance
