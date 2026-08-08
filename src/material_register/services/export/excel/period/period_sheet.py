from pathlib import Path

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, Side, Border
from openpyxl.utils import get_column_letter

from material_register.domain.export_dataclass import ExportItemIn, ExportItemOut, PeriodItemIn, PeriodItemOut
from material_register.services.export.period_report import PeriodReport
from material_register.ui.helpers.formating_utils import format_date_range_to_locale


# noinspection PyDunderSlots
class PeriodSheet:
    START_ROW = 1
    LAST_COLUMN = 8
    TITLE_FONT_SIZE = 12
    DEFAULT_FONT_SIZE = 10
    TITLE_ROW_HEIGHT = 20
    DEFAULT_ROW_HEIGHT = 15
    NOTES_ROWS = 4
    ERROR_TEXT = "[N/A]"

    @staticmethod
    def create_sheet(sheet: Worksheet, export_settings: dict[str, Path | str | float | bool],
                     export_texts: dict[str, str],
                     data_in: list[ExportItemIn], out_data: list[ExportItemOut]) -> tuple[Worksheet, float]:
        period_in_data = PeriodReport.get_period_data_in(data_in)
        period_out_data = PeriodReport.get_period_data_out(out_data)
        row = PeriodSheet.START_ROW
        row = PeriodSheet._create_header(sheet, row, PeriodSheet.LAST_COLUMN, export_settings, export_texts)
        row, buyback_cell, total_value = PeriodSheet._create_financial_section(sheet, row, PeriodSheet.LAST_COLUMN, export_settings,
                                                    export_texts)
        freeze_row = row
        data_section_row = row
        in_section_row, buyback_value = PeriodSheet._create_data_in_section(sheet, data_section_row, PeriodSheet.LAST_COLUMN,
                                                             export_texts, period_in_data)
        total_value = PeriodSheet._update_buyback_value(buyback_cell, buyback_value, total_value)
        out_section_row = PeriodSheet._create_data_out_section(sheet, data_section_row, PeriodSheet.LAST_COLUMN,
                                                               export_texts, period_out_data)
        row = PeriodSheet._create_data_spacer(sheet, in_section_row, out_section_row, PeriodSheet.LAST_COLUMN)
        PeriodSheet._auto_size_columns(sheet)
        page_text = export_texts.get("pageText", PeriodSheet.ERROR_TEXT)
        PeriodSheet._setup_sheet(sheet, row, PeriodSheet.LAST_COLUMN, freeze_row, page_text)
        return sheet, total_value

    @staticmethod
    def _setup_sheet(sheet: Worksheet, last_row: int, last_column: int, freeze_row: int, page_text: str) -> None:
        sheet.sheet_view.zoomScale = 90
        sheet.freeze_panes = f"A{freeze_row}"
        sheet.print_area = f"A1:{get_column_letter(last_column)}{last_row}"
        sheet.print_title_rows = f"1:{freeze_row - 1}"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.print_options.horizontalCentered = True
        sheet.oddFooter.center.text = f"{page_text} &P / &N"

    @staticmethod
    def _create_header(sheet: Worksheet, row: int, last_column: int,
                       export_settings: dict[str, Path | str | float | bool],
                       export_texts: dict[str, str]) -> int:
        start_row = row
        middle_column = last_column // 2
        branch_label_column = middle_column + 1
        branch_value_column = middle_column + 2
        cell = sheet.cell(row=row, column=1, value=export_texts.get("titleText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell)
        PeriodSheet._cell_font(cell, font_size=PeriodSheet.TITLE_FONT_SIZE, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.TITLE_ROW_HEIGHT
        row += 1
        range_text = export_texts.get("rangeText", PeriodSheet.ERROR_TEXT)
        cell = sheet.cell(row=row, column=1, value=range_text)
        PeriodSheet._cell_alignment(cell, horizontal="left")
        PeriodSheet._cell_font(cell, PeriodSheet.DEFAULT_FONT_SIZE, bold=True)
        period_value = PeriodSheet._get_period_range(export_settings.get("from_date", None),
                                                     export_settings.get("to_date", None))
        cell = sheet.cell(row=row, column=2, value=period_value)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=middle_column)
        PeriodSheet._cell_alignment(cell)
        PeriodSheet._cell_font(cell, PeriodSheet.DEFAULT_FONT_SIZE, bold=True)
        branch_text = export_texts.get("branchText", PeriodSheet.ERROR_TEXT)
        cell = sheet.cell(row=row, column=branch_label_column, value=branch_text)
        PeriodSheet._cell_alignment(cell, horizontal="left")
        PeriodSheet._cell_font(cell, PeriodSheet.DEFAULT_FONT_SIZE, bold=True)
        branch_value = export_settings.get("branchNameLineEdit", PeriodSheet.ERROR_TEXT)
        cell = sheet.cell(row=row, column=branch_value_column, value=branch_value)
        sheet.merge_cells(start_row=row, start_column=branch_value_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell)
        PeriodSheet._cell_font(cell, PeriodSheet.DEFAULT_FONT_SIZE, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        PeriodSheet._set_borders(sheet, start_row=start_row, start_column=1, end_row=row, end_column=last_column)
        row += 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _create_financial_section(sheet: Worksheet, row: int, last_column: int,
                                  export_settings: dict[str, Path | str | float | bool],
                                  export_texts: dict[str, str]) -> tuple[int, Cell, float]:
        start_row = row
        middle_column = last_column // 2
        financial_label_column = middle_column + 1
        financial_value_column = middle_column + 3
        currency_suffix = export_texts.get("currencySuffix", PeriodSheet.ERROR_TEXT)
        cell_format = f'#,##0.0 "{currency_suffix}";[Red]#,##0.0 "{currency_suffix}"'
        cell = sheet.cell(row=row, column=1, value=export_texts.get("notesText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=middle_column)
        PeriodSheet._cell_alignment(cell)
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        cell = sheet.cell(row=row, column=financial_label_column,
                          value=export_texts.get("financialText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=financial_label_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell)
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        row += 1
        cell = sheet.cell(row=row, column=1)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row + PeriodSheet.NOTES_ROWS, end_column=middle_column)
        PeriodSheet._cell_alignment(cell, horizontal="left", vertical="top")
        PeriodSheet._cell_font(cell, PeriodSheet.DEFAULT_FONT_SIZE)
        for current_row in range(row, row + PeriodSheet.NOTES_ROWS):
            sheet.row_dimensions[current_row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        opening_balance_row = row
        opening_balance = export_settings.get("openingBalanceSpinbox", 0.0)
        cell = sheet.cell(row=row, column=financial_label_column,
                          value=export_texts.get("openingBalanceText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=financial_label_column, end_row=row,
                          end_column=financial_label_column + 1)
        PeriodSheet._cell_alignment(cell, horizontal="left")
        PeriodSheet._cell_font(cell, bold=True)
        cell = sheet.cell(row=row, column=financial_value_column, value=opening_balance)
        cell.number_format = cell_format
        sheet.merge_cells(start_row=row, start_column=financial_value_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell, horizontal="right")
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        row += 1
        income = export_settings.get("income", 0.0)
        cell = sheet.cell(row=row, column=financial_label_column,
                          value=export_texts.get("incomeText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=financial_label_column, end_row=row,
                          end_column=financial_label_column + 1)
        PeriodSheet._cell_alignment(cell, horizontal="left")
        PeriodSheet._cell_font(cell, bold=True)
        cell = sheet.cell(row=row, column=financial_value_column, value=income)
        cell.number_format = cell_format
        sheet.merge_cells(start_row=row, start_column=financial_value_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell, horizontal="right")
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        row += 1
        buyback_row = row
        cell = sheet.cell(row=row, column=financial_label_column,
                          value=export_texts.get("buybackText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=financial_label_column, end_row=row,
                          end_column=financial_label_column + 1)
        PeriodSheet._cell_alignment(cell, horizontal="left")
        PeriodSheet._cell_font(cell, bold=True)
        cell = sheet.cell(row=row, column=financial_value_column, value=0.0)
        cell.number_format = cell_format
        sheet.merge_cells(start_row=row, start_column=financial_value_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell, horizontal="right")
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        row += 1
        expense_row = row
        expense = export_settings.get("expense", 0.0) * -1
        cell = sheet.cell(row=row, column=financial_label_column,
                          value=export_texts.get("expenseText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=financial_label_column, end_row=row,
                          end_column=financial_label_column + 1)
        PeriodSheet._cell_alignment(cell, horizontal="left")
        PeriodSheet._cell_font(cell, bold=True)
        cell = sheet.cell(row=row, column=financial_value_column, value=expense)
        cell.number_format = cell_format
        sheet.merge_cells(start_row=row, start_column=financial_value_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell, horizontal="right")
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        row += 1
        cell = sheet.cell(row=row, column=financial_label_column,
                          value=export_texts.get("balanceText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=financial_label_column, end_row=row,
                          end_column=financial_label_column + 1)
        PeriodSheet._cell_alignment(cell, horizontal="left")
        PeriodSheet._cell_font(cell, bold=True)
        column_letter = get_column_letter(financial_value_column)
        balance_value = f"=SUM({column_letter}{opening_balance_row}:{column_letter}{expense_row})"
        cell = sheet.cell(row=row, column=financial_value_column)
        cell.value = balance_value
        cell.number_format = cell_format
        sheet.merge_cells(start_row=row, start_column=financial_value_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell, horizontal="right")
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        PeriodSheet._set_borders(sheet, start_row=start_row, start_column=1, end_row=row, end_column=last_column)
        PeriodSheet._set_borders(sheet, start_row=row, start_column=financial_value_column,
                                 end_row=row, end_column=last_column, style="medium")
        row += 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        total = round(opening_balance + income + expense, 1)
        return row + 1, sheet.cell(row=buyback_row, column=financial_value_column), total

    @staticmethod
    def _create_data_in_section(sheet: Worksheet, row: int, last_column: int, export_texts: dict[str, str],
                                in_data: dict[str, list[PeriodItemIn]]) -> tuple[int, float]:
        start_row = row
        first_column = 1
        middle_column = last_column // 2
        currency_suffix = export_texts.get("currencySuffix", PeriodSheet.ERROR_TEXT)
        money_cell_format = f'#,##0.0 "{currency_suffix}";[Red]#,##0.0 "{currency_suffix}"'
        buyback = 0
        cell = sheet.cell(row=row, column=first_column, value=export_texts.get("buybackText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=first_column, end_row=row, end_column=middle_column)
        PeriodSheet._cell_alignment(cell)
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        row += 1
        for category, category_data in in_data.items():
            cell = sheet.cell(row=row, column=first_column, value=category)
            sheet.merge_cells(start_row=row, start_column=first_column, end_row=row, end_column=middle_column)
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
            row += 1
            cell = sheet.cell(row=row, column=first_column,
                              value=export_texts.get("commodityText", PeriodSheet.ERROR_TEXT))
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            cell = sheet.cell(row=row, column=first_column + 1,
                              value=export_texts.get("pricePerUnitText", PeriodSheet.ERROR_TEXT))
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            cell = sheet.cell(row=row, column=first_column + 2,
                              value=export_texts.get("quantityText", PeriodSheet.ERROR_TEXT))
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            cell = sheet.cell(row=row, column=first_column + 3,
                              value=export_texts.get("totalPriceText", PeriodSheet.ERROR_TEXT))
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
            row += 1
            row, category_buyback = PeriodSheet._create_category_in_section(
                sheet, row, last_column, export_texts, category_data, money_cell_format)
            buyback += category_buyback
        PeriodSheet._set_borders(sheet, start_row=start_row, start_column=1, end_row=row - 1, end_column=last_column // 2)
        return row, buyback

    @staticmethod
    def _create_data_out_section(sheet: Worksheet, row: int, last_column: int, export_texts: dict[str, str],
                                 data_out: dict[str, list[PeriodItemOut]]) -> int:
        start_row = row
        middle_column = last_column // 2
        first_column = middle_column + 1
        cell = sheet.cell(row=row, column=first_column, value=export_texts.get("exportText", PeriodSheet.ERROR_TEXT))
        sheet.merge_cells(start_row=row, start_column=first_column, end_row=row, end_column=last_column)
        PeriodSheet._cell_alignment(cell)
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        row += 1
        for category, category_data in data_out.items():
            cell = sheet.cell(row=row, column=first_column, value=category)
            sheet.merge_cells(start_row=row, start_column=first_column, end_row=row, end_column=last_column)
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
            row += 1
            cell = sheet.cell(row=row, column=first_column,
                              value=export_texts.get("commodityText", PeriodSheet.ERROR_TEXT))
            sheet.merge_cells(start_row=row, start_column=first_column, end_row=row, end_column=first_column + 1)
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            cell = sheet.cell(row=row, column=first_column + 2,
                              value=export_texts.get("quantityText", PeriodSheet.ERROR_TEXT))
            sheet.merge_cells(start_row=row, start_column=first_column + 2, end_row=row, end_column=last_column)
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
            row += 1
            row = PeriodSheet._create_category_out_section(sheet, row, last_column, category_data)
        PeriodSheet._set_borders(sheet, start_row=start_row, start_column=first_column, end_row=row - 1, end_column=last_column)
        return row

    @staticmethod
    def _create_category_in_section(sheet: Worksheet, row: int, last_column: int, export_texts: dict[str, str],
                                    in_data: list[PeriodItemIn], money_cell_format: str) -> tuple[int, float]:
        first_column = 1
        middle_column = last_column // 2
        commodity_column = first_column
        price_column = first_column + 1
        quantity_column = first_column + 2
        total_column = first_column + 3
        summary_label_column = middle_column - 1
        buyback = 0.0
        for item in in_data:
            quantity_cell_format = f'#,##0.0 "{item.commodity_unit}";[Red]#,##0.0 "{item.commodity_unit}"'
            cell = sheet.cell(row=row, column=commodity_column, value=item.commodity_name)
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            cell = sheet.cell(row=row, column=price_column, value=item.price_per_unit)
            cell.number_format = money_cell_format
            PeriodSheet._cell_alignment(cell, horizontal="right")
            PeriodSheet._cell_font(cell)
            cell = sheet.cell(row=row, column=quantity_column, value=item.total_quantity)
            cell.number_format = quantity_cell_format
            PeriodSheet._cell_alignment(cell, horizontal="right")
            PeriodSheet._cell_font(cell)
            cell = sheet.cell(row=row, column=total_column, value=item.total_price)
            cell.number_format = money_cell_format
            PeriodSheet._cell_alignment(cell, horizontal="right")
            PeriodSheet._cell_font(cell, bold=True)
            sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
            row += 1
            buyback += item.total_price
        summary_value_column = middle_column
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=middle_column - 2)
        cell = sheet.cell(row=row, column=summary_label_column,
                          value=export_texts.get("summaryPriceText", PeriodSheet.ERROR_TEXT))
        PeriodSheet._cell_alignment(cell, horizontal="right")
        PeriodSheet._cell_font(cell, bold=True)
        cell = sheet.cell(row=row, column=summary_value_column, value=buyback)
        cell.number_format = money_cell_format
        PeriodSheet._cell_alignment(cell, horizontal="right")
        PeriodSheet._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        return row + 1, buyback

    @staticmethod
    def _create_category_out_section(sheet: Worksheet, row: int, last_column: int, out_data: list[PeriodItemOut]) -> int:
        middle_column = last_column // 2
        first_column = middle_column + 1
        commodity_column = first_column
        quantity_column = first_column + 2
        stop_index = len(out_data)
        for index, item in enumerate(out_data):
            quantity_cell_format = f'#,##0.0 "{item.commodity_unit}";[Red]#,##0.0 "{item.commodity_unit}"'
            cell = sheet.cell(row=row, column=commodity_column, value=item.commodity_name)
            sheet.merge_cells(start_row=row, start_column=commodity_column, end_row=row, end_column=commodity_column + 1)
            PeriodSheet._cell_alignment(cell)
            PeriodSheet._cell_font(cell, bold=True)
            cell = sheet.cell(row=row, column=quantity_column, value=item.total_quantity)
            cell.number_format = quantity_cell_format
            sheet.merge_cells(start_row=row, start_column=quantity_column, end_row=row, end_column=last_column)
            PeriodSheet._cell_alignment(cell, horizontal="right")
            PeriodSheet._cell_font(cell, bold=True)
            sheet.row_dimensions[row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
            if index + 1 != stop_index:
                row += 1
        return row + 1

    @staticmethod
    def _create_data_spacer(sheet: Worksheet, in_row: int, out_row: int, last_column: int) -> int:
        middle_column = last_column // 2
        different = abs(in_row - out_row)
        if in_row < out_row:
            last_row = out_row
            sheet.merge_cells(start_row=in_row, start_column=1, end_row=in_row + different, end_column=middle_column)
        else:
            last_row = in_row
            sheet.merge_cells(start_row=out_row, start_column=middle_column + 1, end_row=out_row + different,
                              end_column=last_column)
        sheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=last_column)
        sheet.row_dimensions[last_row].height = PeriodSheet.DEFAULT_ROW_HEIGHT
        return last_row + 1

    @staticmethod
    def _cell_alignment(cell, horizontal: str = "center", vertical: str = "center") -> None:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)

    @staticmethod
    def _cell_font(cell, font_size: int | None = None, bold: bool = False) -> None:
        if font_size is None:
            font_size = PeriodSheet.DEFAULT_FONT_SIZE
        cell.font = Font(size=font_size, bold=bold)

    @staticmethod
    def _set_borders(sheet: Worksheet, start_row: int, start_column: int, end_row: int, end_column: int,
                     style: str = "thin") -> None:
        side = Side(border_style=style)
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row,
                                   min_col=start_column, max_col=end_column):
            for cell in row:
                top = None
                bottom = None
                left = None
                right = None
                if cell.row == start_row:
                    top = side
                if cell.row == end_row:
                    bottom = side
                if cell.column == start_column:
                    left = side
                if cell.column == end_column:
                    right = side
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    @staticmethod
    def _auto_size_columns(sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            sheet.column_dimensions[column_letter].width = max_length + 3

    @staticmethod
    def _update_buyback_value(buyback_cell: Cell, buyback_value: float, total_value: float) -> float:
        buyback_cell.value = buyback_value * -1
        return total_value + buyback_cell.value

    @staticmethod
    def _get_period_range(from_date: str | None, to_date: str | None) -> str:
        if from_date is None or to_date is None:
            return PeriodSheet.ERROR_TEXT
        return format_date_range_to_locale(from_date, to_date)