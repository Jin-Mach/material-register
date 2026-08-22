from pathlib import Path

from openpyxl import Workbook

from material_register.domain.export_dataclass.summary_dataclass import (
    SummaryExportItemIn,
    SummaryExportItemOut,
)
from material_register.services.export.excel.summary_export.summary_sheet import (
    SummarySheet,
)


class SummaryWorkbook:
    ERROR_TEXT = "N/A"

    @staticmethod
    def create_workbook(
        export_settings: dict[str, Path | str | float | bool],
        export_texts: dict[str, dict[str, str]],
        in_data: list[SummaryExportItemIn],
        out_data: list[SummaryExportItemOut],
    ) -> tuple[Workbook, float]:
        workbook = Workbook()
        workbook.remove(workbook.active)
        summary_texts = export_texts.get("SummarySheet", {})
        sheet = workbook.create_sheet(
            summary_texts.get("sheetName", SummaryWorkbook.ERROR_TEXT)
        )
        _, last_balance = SummarySheet.create_sheet(
            sheet, export_settings, summary_texts, in_data, out_data
        )
        return workbook, last_balance
