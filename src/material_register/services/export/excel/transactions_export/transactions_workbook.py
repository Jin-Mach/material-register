from pathlib import Path

from openpyxl import Workbook

from material_register.config.ui_constants import TRANSFER_IN, TRANSFER_OUT
from material_register.domain.export_dataclass.transactions_dataclass import (
    TransactionsExportDay,
)
from material_register.services.export.excel.transactions_export.transactions_day_sheet_in import (
    TransactionsDaySheetIn,
)
from material_register.services.export.excel.transactions_export.transactions_day_sheet_out import (
    TransactionsDaySheetOut,
)
from material_register.utils.date_filters import parse_date


class TransactionsWorkbook:
    ERROR_TEXT = ["N/A"]

    @staticmethod
    def create_workbook(
        export_settings: dict[str, Path | str | float | bool],
        transactions_texts: dict[str, str],
        data: list[TransactionsExportDay],
        transfer_type: str,
    ) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for day_data in data:
            sheet_date = parse_date(day_data.transaction_date)
            sheet_title = sheet_date.strftime("%d.%m.%Y")
            sheet = workbook.create_sheet(sheet_title)
            if transfer_type == TRANSFER_IN:
                TransactionsDaySheetIn.create_sheet(
                    sheet, export_settings, transactions_texts, day_data
                )
            elif transfer_type == TRANSFER_OUT:
                TransactionsDaySheetOut.create_sheet(
                    sheet, export_settings, transactions_texts, day_data
                )
        return workbook
