from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from material_register.domain.export_dataclass.transactions_dataclass import (
    TransactionExportItem,
    TransactionsExportDay,
    TransactionsExportTransaction,
)
from material_register.services.export.excel.excel_helpers import (
    cell_alignment,
    cell_font,
    set_borders,
)
from material_register.ui.helpers.formating_utils import (
    format_date_to_locale,
    format_time_to_locale,
)


# noinspection PyDunderSlots
class TransactionsDaySheetOut:
    START_ROW = 1
    LAST_COLUMN = 6
    TITLE_FONT_SIZE = 12
    DEFAULT_FONT_SIZE = 10
    TITLE_ROW_HEIGHT = 20
    DEFAULT_ROW_HEIGHT = 15
    NOTES_ROWS = 4
    ERROR_TEXT = "[N/A]"

    @staticmethod
    def create_sheet(
        sheet: Worksheet,
        export_settings: dict[str, Path | str | float | bool],
        export_texts: dict[str, str],
        day_data: TransactionsExportDay,
    ) -> Worksheet:
        row = TransactionsDaySheetOut.START_ROW
        row = TransactionsDaySheetOut._create_header(
            sheet,
            row,
            TransactionsDaySheetOut.LAST_COLUMN,
            export_settings,
            export_texts,
            day_data,
        )
        row = TransactionsDaySheetOut._create_count_section(
            sheet,
            row,
            TransactionsDaySheetOut.LAST_COLUMN,
            export_texts,
            day_data,
        )
        freeze_row = row
        row = TransactionsDaySheetOut._create_transactions_section(
            sheet,
            row,
            TransactionsDaySheetOut.LAST_COLUMN,
            export_texts,
            day_data,
        )
        TransactionsDaySheetOut._auto_size_columns(sheet)
        page_text = export_texts.get("pageText", TransactionsDaySheetOut.ERROR_TEXT)
        TransactionsDaySheetOut._setup_sheet(
            sheet,
            row,
            TransactionsDaySheetOut.LAST_COLUMN,
            freeze_row,
            page_text,
        )
        return sheet

    @staticmethod
    def _setup_sheet(
        sheet: Worksheet,
        last_row: int,
        last_column: int,
        freeze_row: int,
        page_text: str,
    ) -> None:
        sheet.sheet_view.zoomScale = 90
        sheet.freeze_panes = f"A{freeze_row}"
        sheet.print_area = f"A1:{get_column_letter(last_column)}{last_row}"
        sheet.print_title_rows = f"1:{freeze_row - 1}"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.print_options.horizontalCentered = True
        sheet.oddFooter.center.text = f"{page_text} &P / &N"

    @staticmethod
    def _create_header(
        sheet: Worksheet,
        row: int,
        last_column: int,
        export_settings: dict[str, Path | str | float | bool],
        export_texts: dict[str, str],
        day_data: TransactionsExportDay,
    ) -> int:
        start_row = row
        middle_column = 4
        cell = sheet.cell(
            row=row,
            column=1,
            value=export_texts.get(
                "titleText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(
            cell,
            font_size=TransactionsDaySheetOut.TITLE_FONT_SIZE,
            bold=True,
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetOut.TITLE_ROW_HEIGHT
        row += 1
        range_text = export_texts.get(
            "rangeText",
            TransactionsDaySheetOut.ERROR_TEXT,
        )
        cell = sheet.cell(row=row, column=1, value=range_text)
        TransactionsDaySheetOut._cell_alignment(
            cell,
            horizontal="left",
        )
        TransactionsDaySheetOut._cell_font(
            cell,
            TransactionsDaySheetOut.DEFAULT_FONT_SIZE,
            bold=True,
        )
        period_value = format_date_to_locale(f"{day_data.transaction_date} 00:00:00")
        cell = sheet.cell(row=row, column=2, value=period_value)
        sheet.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=middle_column - 1,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(
            cell,
            TransactionsDaySheetOut.DEFAULT_FONT_SIZE,
            bold=True,
        )
        branch_text = export_texts.get(
            "branchText",
            TransactionsDaySheetOut.ERROR_TEXT,
        )
        cell = sheet.cell(
            row=row,
            column=middle_column,
            value=branch_text,
        )
        TransactionsDaySheetOut._cell_alignment(
            cell,
            horizontal="left",
        )
        TransactionsDaySheetOut._cell_font(
            cell,
            TransactionsDaySheetOut.DEFAULT_FONT_SIZE,
            bold=True,
        )
        branch_value = export_settings.get(
            "branchNameLineEdit",
            TransactionsDaySheetOut.ERROR_TEXT,
        )
        cell = sheet.cell(
            row=row,
            column=middle_column + 1,
            value=branch_value,
        )
        sheet.merge_cells(
            start_row=row,
            start_column=middle_column + 1,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(
            cell,
            TransactionsDaySheetOut.DEFAULT_FONT_SIZE,
            bold=True,
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        TransactionsDaySheetOut._set_borders(
            sheet,
            start_row=start_row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        row += 1
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _create_count_section(
        sheet: Worksheet,
        row: int,
        last_column: int,
        export_texts: dict[str, str],
        day_data: TransactionsExportDay,
    ) -> int:
        start_row = row
        middle_column = last_column // 2
        count_label_column = middle_column + 1
        count_value_column = middle_column + 3
        category_totals: dict[str, tuple[float, str]] = {}
        for transaction in day_data.transactions_list or []:
            for item in transaction.transaction_items or []:
                category = item.category or TransactionsDaySheetOut.ERROR_TEXT
                quantity = item.unit_count or 0.0
                unit = item.commodity_unit or TransactionsDaySheetOut.ERROR_TEXT
                if category in category_totals:
                    current_quantity, current_unit = category_totals[category]
                    category_totals[category] = (
                        current_quantity + quantity,
                        current_unit,
                    )
                else:
                    category_totals[category] = (
                        quantity,
                        unit,
                    )
        cell = sheet.cell(
            row=row,
            column=1,
            value=export_texts.get(
                "notesText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=middle_column,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(cell, bold=True)
        cell = sheet.cell(
            row=row,
            column=count_label_column,
            value=export_texts.get(
                "countText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=count_label_column,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        row += 1
        notes_rows = max(
            TransactionsDaySheetOut.NOTES_ROWS,
            len(category_totals),
        )
        notes_start_row = row
        notes_end_row = row + notes_rows - 1
        cell = sheet.cell(row=row, column=1)
        sheet.merge_cells(
            start_row=notes_start_row,
            start_column=1,
            end_row=notes_end_row,
            end_column=middle_column,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(
            cell,
            TransactionsDaySheetOut.DEFAULT_FONT_SIZE,
        )
        for current_row in range(notes_start_row, notes_end_row + 1):
            sheet.row_dimensions[
                current_row
            ].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        category_row = row
        for category, (quantity, unit) in category_totals.items():
            cell = sheet.cell(
                row=category_row,
                column=count_label_column,
                value=category,
            )
            sheet.merge_cells(
                start_row=category_row,
                start_column=count_label_column,
                end_row=category_row,
                end_column=count_label_column + 1,
            )
            TransactionsDaySheetOut._cell_alignment(cell)
            TransactionsDaySheetOut._cell_font(cell, bold=True)
            cell = sheet.cell(
                row=category_row,
                column=count_value_column,
                value=quantity,
            )
            cell.number_format = f'#,##0.0 "{unit}";[Red]#,##0.0 "{unit}"'
            sheet.merge_cells(
                start_row=category_row,
                start_column=count_value_column,
                end_row=category_row,
                end_column=last_column,
            )
            TransactionsDaySheetOut._cell_alignment(
                cell,
                horizontal="right",
            )
            TransactionsDaySheetOut._cell_font(cell, bold=True)
            sheet.row_dimensions[
                category_row
            ].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
            category_row += 1
        if category_row <= notes_end_row:
            sheet.merge_cells(
                start_row=category_row,
                start_column=count_label_column,
                end_row=notes_end_row,
                end_column=last_column,
            )
        last_section_row = max(
            notes_end_row,
            category_row - 1,
        )
        TransactionsDaySheetOut._set_borders(
            sheet,
            start_row=start_row,
            start_column=1,
            end_row=last_section_row,
            end_column=last_column,
        )
        spacer_row = last_section_row + 1
        sheet.merge_cells(
            start_row=spacer_row,
            start_column=1,
            end_row=spacer_row,
            end_column=last_column,
        )
        sheet.row_dimensions[
            spacer_row
        ].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        return spacer_row + 1

    @staticmethod
    def _create_transactions_section(
        sheet: Worksheet,
        row: int,
        last_column: int,
        export_texts: dict[str, str],
        day_data: TransactionsExportDay,
    ) -> int:
        if day_data.transactions_list:
            for index, transaction in enumerate(day_data.transactions_list):
                row = TransactionsDaySheetOut._create_transaction(
                    sheet,
                    row,
                    last_column,
                    export_texts,
                    transaction,
                )
                if index + 1 < len(day_data.transactions_list):
                    row = TransactionsDaySheetOut._create_spacer(
                        sheet,
                        row,
                        last_column,
                    )
        return row

    @staticmethod
    def _create_transaction(
        sheet: Worksheet,
        row: int,
        last_column: int,
        export_texts: dict[str, str],
        transaction: TransactionsExportTransaction,
    ) -> int:
        customer_start_column = 1
        customer_end_column = 3
        items_start_column = 4
        items_end_column = last_column
        transaction_start_row = row
        customer_lines = [
            (
                export_texts.get(
                    "documentNumberText",
                    TransactionsDaySheetOut.ERROR_TEXT,
                ),
                transaction.document_number or TransactionsDaySheetOut.ERROR_TEXT,
            ),
            (
                export_texts.get(
                    "customerNameText",
                    TransactionsDaySheetOut.ERROR_TEXT,
                ),
                transaction.customer_name or TransactionsDaySheetOut.ERROR_TEXT,
            ),
            (
                export_texts.get(
                    "addressText",
                    TransactionsDaySheetOut.ERROR_TEXT,
                ),
                transaction.address or TransactionsDaySheetOut.ERROR_TEXT,
            ),
            (
                export_texts.get(
                    "createdAtText",
                    TransactionsDaySheetOut.ERROR_TEXT,
                ),
                format_time_to_locale(transaction.created_at)
                if transaction.created_at
                else TransactionsDaySheetOut.ERROR_TEXT,
            ),
        ]
        row = TransactionsDaySheetOut._create_items_header(
            sheet,
            row,
            export_texts,
            customer_start_column,
            customer_end_column,
            items_start_column,
        )
        customer_data_start_row = row
        for index, (label, value) in enumerate(customer_lines):
            customer_row = customer_data_start_row + index
            cell = sheet.cell(
                row=customer_row,
                column=customer_start_column,
                value=f"{label} {value}",
            )
            sheet.merge_cells(
                start_row=customer_row,
                start_column=customer_start_column,
                end_row=customer_row,
                end_column=customer_end_column,
            )
            TransactionsDaySheetOut._cell_alignment(
                cell,
                horizontal="left",
            )
            TransactionsDaySheetOut._cell_font(cell)
            sheet.row_dimensions[
                customer_row
            ].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        items_row = row
        item_quantity_cells = []
        if transaction.transaction_items:
            for item in transaction.transaction_items:
                item_row = items_row
                items_row = TransactionsDaySheetOut._create_item(
                    sheet,
                    items_row,
                    item,
                    items_start_column,
                )
                if item.unit_count is not None:
                    item_quantity_cells.append(
                        sheet.cell(
                            row=item_row,
                            column=items_start_column + 2,
                        )
                    )
        total_row = max(
            items_row,
            customer_data_start_row + len(customer_lines) - 1,
        )
        cell = sheet.cell(
            row=total_row,
            column=items_end_column - 1,
            value=export_texts.get(
                "summaryPriceText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
        )
        TransactionsDaySheetOut._cell_alignment(
            cell,
            horizontal="right",
        )
        TransactionsDaySheetOut._cell_font(cell, bold=True)
        transaction_total_cell = sheet.cell(
            row=total_row,
            column=items_end_column,
        )
        if item_quantity_cells:
            transaction_total_cell.value = (
                f"=SUM({','.join(cell.coordinate for cell in item_quantity_cells)})"
            )
        else:
            transaction_total_cell.value = "=0"
        TransactionsDaySheetOut._cell_alignment(
            transaction_total_cell,
            horizontal="right",
        )
        TransactionsDaySheetOut._cell_font(
            transaction_total_cell,
            bold=True,
        )
        TransactionsDaySheetOut._set_borders(
            sheet,
            start_row=total_row,
            start_column=items_end_column,
            end_row=total_row,
            end_column=items_end_column,
            style="medium",
        )
        sheet.row_dimensions[
            total_row
        ].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        TransactionsDaySheetOut._set_borders(
            sheet,
            start_row=transaction_start_row,
            start_column=customer_start_column,
            end_row=total_row,
            end_column=last_column,
        )
        TransactionsDaySheetOut._set_borders(
            sheet,
            start_row=total_row,
            start_column=items_end_column,
            end_row=total_row,
            end_column=items_end_column,
            style="medium",
        )
        return total_row + 1

    @staticmethod
    def _create_items_header(
        sheet: Worksheet,
        row: int,
        export_texts: dict[str, str],
        customer_start_column: int,
        customer_end_column: int,
        start_column: int,
    ) -> int:
        cell = sheet.cell(
            row=row,
            column=customer_start_column,
            value=export_texts.get(
                "customerNameText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=customer_start_column,
            end_row=row,
            end_column=customer_end_column,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(cell, bold=True)
        headers = [
            export_texts.get(
                "categoryText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
            export_texts.get(
                "commodityText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
            export_texts.get(
                "quantityText",
                TransactionsDaySheetOut.ERROR_TEXT,
            ),
        ]
        for column, value in zip(
            range(start_column, start_column + len(headers)),
            headers,
        ):
            cell = sheet.cell(
                row=row,
                column=column,
                value=value,
            )
            TransactionsDaySheetOut._cell_alignment(cell)
            TransactionsDaySheetOut._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _create_item(
        sheet: Worksheet,
        row: int,
        item: TransactionExportItem,
        start_column: int,
    ) -> int:
        quantity_cell_format = (
            f'#,##0.0 "{item.commodity_unit}";[Red]#,##0.0 "{item.commodity_unit}"'
        )
        cell = sheet.cell(
            row=row,
            column=start_column,
            value=item.category,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(cell)
        cell = sheet.cell(
            row=row,
            column=start_column + 1,
            value=item.commodity_name,
        )
        TransactionsDaySheetOut._cell_alignment(cell)
        TransactionsDaySheetOut._cell_font(cell)
        cell = sheet.cell(
            row=row,
            column=start_column + 2,
            value=item.unit_count,
        )
        cell.number_format = quantity_cell_format
        TransactionsDaySheetOut._cell_alignment(
            cell,
            horizontal="right",
        )
        TransactionsDaySheetOut._cell_font(cell)
        sheet.row_dimensions[row].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _create_spacer(
        sheet: Worksheet,
        row: int,
        last_column: int,
    ) -> int:
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetOut.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _cell_alignment(
        cell,
        horizontal: str = "center",
        vertical: str = "center",
    ) -> None:
        cell_alignment(cell, horizontal=horizontal, vertical=vertical)

    @staticmethod
    def _cell_font(
        cell,
        font_size: int | None = None,
        bold: bool = False,
    ) -> None:
        cell_font(
            cell,
            font_size=font_size,
            bold=bold,
            default_font_size=TransactionsDaySheetOut.DEFAULT_FONT_SIZE,
        )

    @staticmethod
    def _set_borders(
        sheet: Worksheet,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
        style: str = "thin",
    ) -> None:
        set_borders(sheet, start_row, start_column, end_row, end_column, style=style)

    @staticmethod
    def _auto_size_columns(sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    length = len(str(cell.value))
                    max_length = max(max_length, length)
            sheet.column_dimensions[column_letter].width = max_length + 3
