from pathlib import Path

from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from material_register.domain.export_dataclass.transactions_dataclass import (
    TransactionExportItem,
    TransactionsExportDay,
    TransactionsExportTransaction,
)
from material_register.ui.helpers.formating_utils import (
    format_date_range_to_locale,
    format_time_to_locale,
)


# noinspection PyDunderSlots
class TransactionsDaySheetIn:
    START_ROW = 1
    LAST_COLUMN = 8
    TITLE_FONT_SIZE = 12
    DEFAULT_FONT_SIZE = 10
    TITLE_ROW_HEIGHT = 20
    DEFAULT_ROW_HEIGHT = 15
    NOTES_ROWS = 3
    ERROR_TEXT = "[N/A]"

    @staticmethod
    def create_sheet(
        sheet: Worksheet,
        export_settings: dict[str, Path | str | float | bool],
        export_texts: dict[str, str],
        day_data: TransactionsExportDay,
    ) -> Worksheet:
        row = TransactionsDaySheetIn.START_ROW
        row = TransactionsDaySheetIn._create_header(
            sheet,
            row,
            TransactionsDaySheetIn.LAST_COLUMN,
            export_settings,
            export_texts,
        )
        cash_cell, transfer_cell, total_cell, row = (
            TransactionsDaySheetIn._create_financial_section(
                sheet, row, TransactionsDaySheetIn.LAST_COLUMN, export_texts
            )
        )
        freeze_row = row
        row, transaction_total_cells = (
            TransactionsDaySheetIn._create_transactions_section(
                sheet, row, TransactionsDaySheetIn.LAST_COLUMN, export_texts, day_data
            )
        )
        TransactionsDaySheetIn._set_financial_formulas(
            cash_cell, transfer_cell, total_cell, transaction_total_cells
        )
        TransactionsDaySheetIn._auto_size_columns(sheet)
        page_text = export_texts.get("pageText", TransactionsDaySheetIn.ERROR_TEXT)
        TransactionsDaySheetIn._setup_sheet(
            sheet, row, TransactionsDaySheetIn.LAST_COLUMN, freeze_row, page_text
        )
        return sheet

    @staticmethod
    def _setup_sheet(
        sheet: Worksheet,
        last_row: int,
        last_column: int,
        freeze_row: int,
        page_text: str,
    ) -> None:
        sheet.sheet_view.zoomScale = 90
        sheet.freeze_panes = f"A{freeze_row}"
        sheet.print_area = f"A1:{get_column_letter(last_column)}{last_row}"
        sheet.print_title_rows = f"1:{freeze_row - 1}"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.print_options.horizontalCentered = True
        sheet.oddFooter.center.text = f"{page_text} &P / &N"

    @staticmethod
    def _create_header(
        sheet: Worksheet,
        row: int,
        last_column: int,
        export_settings: dict[str, Path | str | float | bool],
        export_texts: dict[str, str],
    ) -> int:
        start_row = row
        middle_column = 4
        cell = sheet.cell(
            row=row,
            column=1,
            value=export_texts.get("titleText", TransactionsDaySheetIn.ERROR_TEXT),
        )
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=last_column
        )
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(
            cell, font_size=TransactionsDaySheetIn.TITLE_FONT_SIZE, bold=True
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.TITLE_ROW_HEIGHT
        row += 1
        range_text = export_texts.get("rangeText", TransactionsDaySheetIn.ERROR_TEXT)
        cell = sheet.cell(row=row, column=1, value=range_text)
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="left")
        TransactionsDaySheetIn._cell_font(
            cell, TransactionsDaySheetIn.DEFAULT_FONT_SIZE, bold=True
        )
        period_value = TransactionsDaySheetIn._get_period_range(
            export_settings.get("from_date", None), export_settings.get("to_date", None)
        )
        cell = sheet.cell(row=row, column=2, value=period_value)
        sheet.merge_cells(
            start_row=row, start_column=2, end_row=row, end_column=middle_column - 1
        )
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(
            cell, TransactionsDaySheetIn.DEFAULT_FONT_SIZE, bold=True
        )
        branch_text = export_texts.get("branchText", TransactionsDaySheetIn.ERROR_TEXT)
        cell = sheet.cell(row=row, column=middle_column, value=branch_text)
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="left")
        TransactionsDaySheetIn._cell_font(
            cell, TransactionsDaySheetIn.DEFAULT_FONT_SIZE, bold=True
        )
        branch_value = export_settings.get(
            "branchNameLineEdit", TransactionsDaySheetIn.ERROR_TEXT
        )
        cell = sheet.cell(row=row, column=middle_column + 1, value=branch_value)
        sheet.merge_cells(
            start_row=row,
            start_column=middle_column + 1,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(
            cell, TransactionsDaySheetIn.DEFAULT_FONT_SIZE, bold=True
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        TransactionsDaySheetIn._set_borders(
            sheet,
            start_row=start_row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        row += 1
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=last_column
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _create_financial_section(
        sheet: Worksheet, row: int, last_column: int, export_texts: dict[str, str]
    ) -> tuple[Cell, Cell, Cell, int]:
        start_row = row
        middle_column = last_column // 2
        financial_label_column = middle_column + 1
        financial_value_column = middle_column + 3
        currency_suffix = export_texts.get(
            "currencySuffix", TransactionsDaySheetIn.ERROR_TEXT
        )
        cell_format = f'#,##0.0 "{currency_suffix}";[Red]#,##0.0 "{currency_suffix}"'
        cell = sheet.cell(
            row=row,
            column=1,
            value=export_texts.get("notesText", TransactionsDaySheetIn.ERROR_TEXT),
        )
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=middle_column
        )
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        cell = sheet.cell(
            row=row,
            column=financial_label_column,
            value=export_texts.get("financialText", TransactionsDaySheetIn.ERROR_TEXT),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=financial_label_column,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        row += 1
        cell = sheet.cell(row=row, column=1)
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row + TransactionsDaySheetIn.NOTES_ROWS,
            end_column=middle_column,
        )
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="left", vertical="top")
        TransactionsDaySheetIn._cell_font(
            cell, TransactionsDaySheetIn.DEFAULT_FONT_SIZE
        )
        for current_row in range(row, row + TransactionsDaySheetIn.NOTES_ROWS):
            sheet.row_dimensions[
                current_row
            ].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        cell = sheet.cell(
            row=row,
            column=financial_label_column,
            value=export_texts.get("CASH", TransactionsDaySheetIn.ERROR_TEXT),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=financial_label_column,
            end_row=row,
            end_column=financial_label_column + 1,
        )
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="left")
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        cash_cell = sheet.cell(row=row, column=financial_value_column)
        cash_cell.number_format = cell_format
        sheet.merge_cells(
            start_row=row,
            start_column=financial_value_column,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetIn._cell_alignment(cash_cell, horizontal="right")
        TransactionsDaySheetIn._cell_font(cash_cell, bold=True)
        row += 1
        cell = sheet.cell(
            row=row,
            column=financial_label_column,
            value=export_texts.get("TRANSFER", TransactionsDaySheetIn.ERROR_TEXT),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=financial_label_column,
            end_row=row,
            end_column=financial_label_column + 1,
        )
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="left")
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        transfer_cell = sheet.cell(row=row, column=financial_value_column)
        transfer_cell.number_format = cell_format
        sheet.merge_cells(
            start_row=row,
            start_column=financial_value_column,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetIn._cell_alignment(transfer_cell, horizontal="right")
        TransactionsDaySheetIn._cell_font(transfer_cell, bold=True)
        row += 1
        cell = sheet.cell(
            row=row,
            column=financial_label_column,
            value=export_texts.get(
                "summaryPriceText", TransactionsDaySheetIn.ERROR_TEXT
            ),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=financial_label_column,
            end_row=row,
            end_column=financial_label_column + 1,
        )
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="left")
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        total_cell = sheet.cell(row=row, column=financial_value_column)
        total_cell.number_format = cell_format
        sheet.merge_cells(
            start_row=row,
            start_column=financial_value_column,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetIn._cell_alignment(total_cell, horizontal="right")
        TransactionsDaySheetIn._cell_font(total_cell, bold=True)
        TransactionsDaySheetIn._set_borders(
            sheet,
            start_row=start_row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        TransactionsDaySheetIn._set_borders(
            sheet,
            start_row=row,
            start_column=financial_value_column,
            end_row=row,
            end_column=last_column,
            style="medium",
        )
        row += 1
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=last_column
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        return cash_cell, transfer_cell, total_cell, row + 1

    @staticmethod
    def _set_financial_formulas(
        cash_cell: Cell,
        transfer_cell: Cell,
        total_cell: Cell,
        transaction_total_cells: list[tuple[str, Cell]],
    ) -> None:
        cash_cells = []
        for payment_type, cell in transaction_total_cells:
            if payment_type == "CASH":
                cash_cells.append(cell.coordinate)
        transfer_cells = []
        for payment_type, cell in transaction_total_cells:
            if payment_type == "TRANSFER":
                transfer_cells.append(cell.coordinate)
        if cash_cells:
            cash_cell.value = f"=SUM({','.join(cash_cells)})"
        else:
            cash_cell.value = "=0"
        if transfer_cells:
            transfer_cell.value = f"=SUM({','.join(transfer_cells)})"
        else:
            transfer_cell.value = "=0"
        total_cell.value = f"=SUM({cash_cell.coordinate},{transfer_cell.coordinate})"

    @staticmethod
    def _create_transactions_section(
        sheet: Worksheet,
        row: int,
        last_column: int,
        export_texts: dict[str, str],
        day_data: TransactionsExportDay,
    ) -> tuple[int, list[tuple[str, Cell]]]:
        transaction_total_cells = []
        if day_data.transactions_list:
            for index, transaction in enumerate(day_data.transactions_list):
                row, transaction_total_cell = (
                    TransactionsDaySheetIn._create_transaction(
                        sheet, row, last_column, export_texts, transaction
                    )
                )
                if transaction_total_cell is not None:
                    transaction_total_cells.append(
                        (transaction.payment_type, transaction_total_cell)
                    )
                if index + 1 < len(day_data.transactions_list):
                    row = TransactionsDaySheetIn._create_spacer(sheet, row, last_column)
        return row, transaction_total_cells

    @staticmethod
    def _create_transaction(
        sheet: Worksheet,
        row: int,
        last_column: int,
        export_texts: dict[str, str],
        transaction: TransactionsExportTransaction,
    ) -> tuple[int, Cell | None]:
        customer_start_column = 1
        customer_end_column = 3
        items_start_column = 4
        items_end_column = last_column
        transaction_start_row = row
        customer_lines = [
            (
                export_texts.get(
                    "documentNumberText", TransactionsDaySheetIn.ERROR_TEXT
                ),
                transaction.document_number or TransactionsDaySheetIn.ERROR_TEXT,
            ),
            (
                export_texts.get("customerNameText", TransactionsDaySheetIn.ERROR_TEXT),
                transaction.customer_name or TransactionsDaySheetIn.ERROR_TEXT,
            ),
            (
                export_texts.get("addressText", TransactionsDaySheetIn.ERROR_TEXT),
                transaction.address or TransactionsDaySheetIn.ERROR_TEXT,
            ),
            (
                export_texts.get("createdAtText", TransactionsDaySheetIn.ERROR_TEXT),
                format_time_to_locale(transaction.created_at)
                if transaction.created_at
                else TransactionsDaySheetIn.ERROR_TEXT,
            ),
            (
                export_texts.get("paymentTypeText", TransactionsDaySheetIn.ERROR_TEXT),
                export_texts.get(
                    transaction.payment_type,
                    transaction.payment_type or TransactionsDaySheetIn.ERROR_TEXT,
                ),
            ),
        ]
        row = TransactionsDaySheetIn._create_items_header(
            sheet,
            row,
            export_texts,
            customer_start_column,
            customer_end_column,
            items_start_column,
            items_end_column,
        )
        customer_data_start_row = row
        for index, (label, value) in enumerate(customer_lines):
            customer_row = customer_data_start_row + index
            cell = sheet.cell(
                row=customer_row, column=customer_start_column, value=f"{label} {value}"
            )
            sheet.merge_cells(
                start_row=customer_row,
                start_column=customer_start_column,
                end_row=customer_row,
                end_column=customer_end_column,
            )
            TransactionsDaySheetIn._cell_alignment(cell, horizontal="left")
            TransactionsDaySheetIn._cell_font(cell)
            sheet.row_dimensions[
                customer_row
            ].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        items_row = row
        item_total_cells = []
        if transaction.transaction_items:
            for item in transaction.transaction_items:
                item_row = items_row
                items_row = TransactionsDaySheetIn._create_item(
                    sheet,
                    items_row,
                    export_texts,
                    item,
                    items_start_column,
                    items_end_column,
                )
                if item.unit_count is not None and item.price_per_unit is not None:
                    item_total_cells.append(
                        sheet.cell(row=item_row, column=items_end_column)
                    )
        total_row = max(items_row, customer_data_start_row + len(customer_lines) - 1)
        cell = sheet.cell(
            row=total_row,
            column=items_end_column - 1,
            value=export_texts.get(
                "summaryPriceText", TransactionsDaySheetIn.ERROR_TEXT
            ),
        )
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="right")
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        currency_suffix = export_texts.get(
            "currencySuffix", TransactionsDaySheetIn.ERROR_TEXT
        )
        transaction_total_cell = sheet.cell(row=total_row, column=items_end_column)
        if item_total_cells:
            transaction_total_cell.value = (
                f"=SUM({','.join(cell.coordinate for cell in item_total_cells)})"
            )
        else:
            transaction_total_cell.value = "=0"
        transaction_total_cell.number_format = (
            f'#,##0.0 "{currency_suffix}";[Red]#,##0.0 "{currency_suffix}"'
        )
        TransactionsDaySheetIn._cell_alignment(
            transaction_total_cell, horizontal="right"
        )
        TransactionsDaySheetIn._cell_font(transaction_total_cell, bold=True)
        TransactionsDaySheetIn._set_borders(
            sheet,
            start_row=total_row,
            start_column=items_end_column,
            end_row=total_row,
            end_column=items_end_column,
            style="medium",
        )
        sheet.row_dimensions[
            total_row
        ].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        TransactionsDaySheetIn._set_borders(
            sheet,
            start_row=transaction_start_row,
            start_column=customer_start_column,
            end_row=total_row,
            end_column=last_column,
        )
        TransactionsDaySheetIn._set_borders(
            sheet,
            start_row=total_row,
            start_column=items_end_column,
            end_row=total_row,
            end_column=items_end_column,
            style="medium",
        )
        return total_row + 1, transaction_total_cell

    @staticmethod
    def _create_items_header(
        sheet: Worksheet,
        row: int,
        export_texts: dict[str, str],
        customer_start_column: int,
        customer_end_column: int,
        start_column: int,
        end_column: int,
    ) -> int:
        cell = sheet.cell(
            row=row,
            column=customer_start_column,
            value=export_texts.get(
                "customerNameText", TransactionsDaySheetIn.ERROR_TEXT
            ),
        )
        sheet.merge_cells(
            start_row=row,
            start_column=customer_start_column,
            end_row=row,
            end_column=customer_end_column,
        )
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        headers = [
            export_texts.get("categoryText", TransactionsDaySheetIn.ERROR_TEXT),
            export_texts.get("commodityText", TransactionsDaySheetIn.ERROR_TEXT),
            export_texts.get("quantityText", TransactionsDaySheetIn.ERROR_TEXT),
            export_texts.get("pricePerUnitText", TransactionsDaySheetIn.ERROR_TEXT),
            export_texts.get("totalPriceText", TransactionsDaySheetIn.ERROR_TEXT),
        ]
        for column, value in zip(range(start_column, end_column + 1), headers):
            cell = sheet.cell(row=row, column=column, value=value)
            TransactionsDaySheetIn._cell_alignment(cell)
            TransactionsDaySheetIn._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _create_item(
        sheet: Worksheet,
        row: int,
        export_texts: dict[str, str],
        item: TransactionExportItem,
        start_column: int,
        end_column: int,
    ) -> int:
        currency_suffix = export_texts.get(
            "currencySuffix", TransactionsDaySheetIn.ERROR_TEXT
        )
        money_cell_format = (
            f'#,##0.0 "{currency_suffix}";[Red]#,##0.0 "{currency_suffix}"'
        )
        quantity_cell_format = (
            f'#,##0.0 "{item.commodity_unit}";[Red]#,##0.0 "{item.commodity_unit}"'
        )
        cell = sheet.cell(row=row, column=start_column, value=item.category)
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(cell)
        cell = sheet.cell(row=row, column=start_column + 1, value=item.commodity_name)
        TransactionsDaySheetIn._cell_alignment(cell)
        TransactionsDaySheetIn._cell_font(cell)
        cell = sheet.cell(row=row, column=start_column + 2, value=item.unit_count)
        cell.number_format = quantity_cell_format
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="right")
        TransactionsDaySheetIn._cell_font(cell)
        cell = sheet.cell(row=row, column=start_column + 3, value=item.price_per_unit)
        cell.number_format = money_cell_format
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="right")
        TransactionsDaySheetIn._cell_font(cell)
        total_price = None
        if item.unit_count is not None and item.price_per_unit is not None:
            total_price = item.unit_count * item.price_per_unit
        cell = sheet.cell(row=row, column=start_column + 4, value=total_price)
        cell.number_format = money_cell_format
        TransactionsDaySheetIn._cell_alignment(cell, horizontal="right")
        TransactionsDaySheetIn._cell_font(cell, bold=True)
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _create_spacer(sheet: Worksheet, row: int, last_column: int) -> int:
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=last_column
        )
        sheet.row_dimensions[row].height = TransactionsDaySheetIn.DEFAULT_ROW_HEIGHT
        return row + 1

    @staticmethod
    def _cell_alignment(
        cell, horizontal: str = "center", vertical: str = "center"
    ) -> None:
        cell.alignment = Alignment(
            horizontal=horizontal, vertical=vertical, wrap_text=True
        )

    @staticmethod
    def _cell_font(cell, font_size: int | None = None, bold: bool = False) -> None:
        if font_size is None:
            font_size = TransactionsDaySheetIn.DEFAULT_FONT_SIZE
        cell.font = Font(size=font_size, bold=bold)

    @staticmethod
    def _set_borders(
        sheet: Worksheet,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
        style: str = "thin",
    ) -> None:
        side = Side(border_style=style)
        for row in sheet.iter_rows(
            min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column
        ):
            for cell in row:
                top = None
                bottom = None
                left = None
                right = None
                if cell.row == start_row:
                    top = side
                if cell.row == end_row:
                    bottom = side
                if cell.column == start_column:
                    left = side
                if cell.column == end_column:
                    right = side
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    @staticmethod
    def _auto_size_columns(sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    length = len(str(cell.value))
                    max_length = max(max_length, length)
            sheet.column_dimensions[column_letter].width = max_length + 3

    @staticmethod
    def _get_period_range(from_date: str | None, to_date: str | None) -> str:
        if from_date is None or to_date is None:
            return TransactionsDaySheetIn.ERROR_TEXT
        return format_date_range_to_locale(from_date, to_date)
